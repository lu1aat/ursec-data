#!/usr/bin/env python3
"""
Extrae texto/metadata de los archivos nuevos en db/files.csv (según diff contra
el último commit de git) para que un agente pueda leerlos y armar un resumen
de la actualización. Usado por el skill de "resumen de actualización".

Uso:
  python3 extract_text.py --new-meta      # JSON con metadata de archivos nuevos
  python3 extract_text.py --new-text      # texto extraído de cada archivo nuevo
  python3 extract_text.py --file RUTA     # extrae texto de un archivo puntual
"""

import argparse
import csv
import io
import json
import subprocess
import sys
from pathlib import Path

DB_FILE   = Path("db/files.csv")
MAX_CHARS = 6000


# ---------------------------------------------------------------------------
# Detección de archivos nuevos (diff contra el último commit)
# ---------------------------------------------------------------------------

def _read_csv_text(text: str) -> dict:
    return {row["url"]: row for row in csv.DictReader(io.StringIO(text))}


def _git_show(rev: str, path: str) -> str | None:
    try:
        out = subprocess.run(
            ["git", "show", f"{rev}:{path}"],
            capture_output=True, text=True, check=True,
        )
        return out.stdout
    except subprocess.CalledProcessError:
        return None


def get_new_rows() -> list[dict]:
    """Filas de db/files.csv ausentes en el último commit.

    Compara primero el working tree vs HEAD (diff sin commitear, el caso
    normal al correr esto justo después de downloader.py). Si no hay
    diferencia (ya se commiteó), compara HEAD vs HEAD~1.
    """
    if not DB_FILE.exists():
        return []
    current = {row["url"]: row for row in csv.DictReader(open(DB_FILE, encoding="utf-8"))}

    head_text = _git_show("HEAD", str(DB_FILE))
    head_rows = _read_csv_text(head_text) if head_text else {}

    new_rows = [row for url, row in current.items() if url not in head_rows]
    if new_rows:
        return new_rows

    prev_text = _git_show("HEAD~1", str(DB_FILE))
    prev_rows = _read_csv_text(prev_text) if prev_text else {}
    return [row for url, row in head_rows.items() if url not in prev_rows]


# ---------------------------------------------------------------------------
# Extracción por tipo de archivo
# ---------------------------------------------------------------------------

def extract_pdf(path: Path) -> str:
    from pypdf import PdfReader
    reader = PdfReader(str(path))
    parts = []
    for page in reader.pages:
        try:
            parts.append(page.extract_text() or "")
        except Exception:
            continue
    return "\n".join(parts)


def extract_docx(path: Path) -> str:
    import docx
    doc = docx.Document(str(path))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def extract_odt(path: Path) -> str:
    from odf.opendocument import load
    from odf.text import P
    from odf.teletype import extractText
    doc = load(str(path))
    return "\n".join(extractText(p) for p in doc.getElementsByType(P))


def extract_xlsx_meta(path: Path) -> str:
    """XLSX es tabular, no narrativo: se listan hojas y encabezados de columna."""
    from openpyxl import load_workbook
    wb = load_workbook(str(path), read_only=True, data_only=True)
    lines = []
    for name in wb.sheetnames:
        ws = wb[name]
        header = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            header = [str(c) for c in row if c is not None]
        lines.append(f"Hoja: {name} ({ws.max_row} filas x {ws.max_column} columnas)")
        if header:
            lines.append("Columnas: " + ", ".join(header))
    wb.close()
    return "\n".join(lines)


def extract_ods_meta(path: Path) -> str:
    """ODS es tabular, no narrativo: se listan hojas y encabezados de columna."""
    from odf.opendocument import load
    from odf.table import Table, TableRow, TableCell
    from odf.teletype import extractText
    doc = load(str(path))
    lines = []
    for table in doc.getElementsByType(Table):
        rows = table.getElementsByType(TableRow)
        header = []
        if rows:
            header = [extractText(c) for c in rows[0].getElementsByType(TableCell) if extractText(c)]
        lines.append(f"Hoja: {table.getAttribute('name')} ({len(rows)} filas)")
        if header:
            lines.append("Columnas: " + ", ".join(header))
    return "\n".join(lines)


EXTRACTORS = {
    "pdf":  extract_pdf,
    "docx": extract_docx,
    "odt":  extract_odt,
    "xlsx": extract_xlsx_meta,
    "ods":  extract_ods_meta,
}


def extract_any(path: Path) -> str:
    ext = path.suffix.lower().lstrip(".")
    extractor = EXTRACTORS.get(ext)
    if not extractor:
        return "(tipo de archivo no soportado para extracción de texto)"

    try:
        text = extractor(path).strip()
    except Exception as exc:
        return f"(error al extraer texto: {exc})"

    if not text:
        return "(sin texto extraíble — probablemente escaneado / imagen)"
    if len(text) > MAX_CHARS:
        text = text[:MAX_CHARS] + f"\n[...truncado, {len(text) - MAX_CHARS} caracteres más...]"
    return text


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def meta_row(r: dict) -> dict:
    fname = r.get("filename", "")
    return {
        "url":            r.get("url", ""),
        "filename":       fname,
        "title":          r.get("title", ""),
        "date_published": r.get("date_published", ""),
        "category":       r.get("category", ""),
        "source":         r.get("source", "datos"),
        "size_str":       r.get("size_str", ""),
        "ext":            fname.rsplit(".", 1)[-1].lower() if "." in fname else "",
        "local_path":     r.get("local_path", ""),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Extrae texto de archivos URSEC para resúmenes")
    parser.add_argument("--new-meta", action="store_true",
                         help="JSON con metadata de archivos nuevos desde el último commit")
    parser.add_argument("--new-text", action="store_true",
                         help="Texto extraído de cada archivo nuevo desde el último commit")
    parser.add_argument("--file", metavar="RUTA",
                         help="Extrae texto de un archivo puntual")
    args = parser.parse_args()

    if args.file:
        print(extract_any(Path(args.file)))
        return 0

    rows = get_new_rows()

    if args.new_meta:
        print(json.dumps([meta_row(r) for r in rows], ensure_ascii=False, indent=2))
        return 0

    if args.new_text:
        if not rows:
            print("No hay archivos nuevos desde el último commit.")
            return 0
        for r in rows:
            m = meta_row(r)
            print("=" * 80)
            print(m["title"] or m["filename"])
            print(f"Archivo: {m['filename']} | Categoría: {m['category'] or '—'} | "
                  f"Fecha: {m['date_published']} | Fuente: {m['source']}")
            print(f"URL: {m['url']}")
            print("-" * 80)
            local_path = m["local_path"]
            if not local_path or not Path(local_path).exists():
                print("(archivo local no encontrado, no se pudo extraer texto)")
            else:
                print(extract_any(Path(local_path)))
            print()
        return 0

    parser.print_help()
    return 0


if __name__ == "__main__":
    sys.exit(main())
